from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
import openpyxl


class BingPilot():
    def __init__(self):
        self.chrome_options = Options()
        self.chrome_options.add_argument("--incognito")

        self.archive = openpyxl.load_workbook('utils/src/accounts.xlsx')

        self.driver = None
        self.wait = None

    def openChrome(self):
        self.driver = webdriver.Chrome(options=self.chrome_options)
        self.driver.maximize_window()

        self.wait = WebDriverWait(self.driver, 10)

    def closeChrome(self):
        if self.driver:
            self.driver.quit()
            self.driver = None
            self.wait = None

    def obtainEmails(self, rMin, rMax):
        spreadsheet = self.archive["Accounts"]

        credentials = {}

        for row in spreadsheet.iter_rows(min_row=rMin, max_row=rMax, min_col=2, max_col=3):
            email = row[0].value
            senha = row[1].value

            credentials[email] = senha
            
        return credentials

    def openBing(self, account, password):
        self.driver.get("https://login.live.com/login.srf?wa=wsignin1.0&rpsnv=151&id=264960&wreply=https%3a%2f%2fwww.bing.com%2fsecure%2fPassport.aspx%3fedge_suppress_profile_switch%3d1%26requrl%3dhttps%253a%252f%252fwww.bing.com%252f%253fcc%253dbr%2526wlexpsignin%253d1%26sig%3d1C60459059DB624322E751E558B263A0%26nopa%3d2&wp=MBI_SSL&lc=1046&CSRFToken=f32eb264-e4cc-4043-a1c6-6a14b52819e6&cobrandid=c333cba8-c15c-4458-b082-7c8ce81bee85&aadredir=1&nopa=2")

        email_field = self.wait.until(EC.visibility_of_element_located((By.ID, "i0116")))
        email_field.send_keys(account)

        self.driver.find_element(By.ID, "idSIButton9").click()

        password_field = self.wait.until(EC.visibility_of_element_located((By.ID, "i0118")))
        password_field.send_keys(password)

        self.driver.find_element(By.ID, "idSIButton9").click()

        try:
            no_button = self.wait.until(EC.element_to_be_clickable((By.ID, "declineButton")))
            no_button.click()
        except TimeoutException:
            pass

        try:
            yes_button = self.wait.until(EC.element_to_be_clickable((By.ID, "bnp_btn_accept")))
            yes_button.click()
        except TimeoutException:
            pass

        self.wait.until(EC.url_contains("https://www.bing.com"))

    def startSearch(self, textToSearch, loops):
        search_field = self.wait.until(EC.visibility_of_element_located((By.ID, "sb_form_q")))
        search_field.send_keys(textToSearch)
        search_field.submit()

        time.sleep(8)

        for _ in range(loops):
            result = self.wait.until(EC.element_to_be_clickable((By.ID, "sb_form_q")))
            result.click()

            for _ in range(2):
                self.driver.find_element(By.ID, "sb_form_q").send_keys(Keys.DELETE)

            self.driver.find_element(By.ID, "sb_form_q").send_keys(Keys.ENTER)

            time.sleep(8)


    def dayStreak(self):
        self.driver.get("https://rewards.bing.com/?ref=rewardspanel")

        elements = self.driver.find_elements(By.CSS_SELECTOR, '.actionLink.x-hidden-vp1')

        time.sleep(1)

        try:
            for link in elements[:3]:
                link.click()
                time.sleep(8)
        except:
            self.driver.find_element(By.CSS_SELECTOR, '[data-bi-id="rx-nav-streak-protection-popup-close"]').click()

            time.sleep(2)

            for link in elements[:3]:
                link.click()
                time.sleep(8)

    def getAccounts(self):
        spreadsheet = self.archive["Accounts"]

        rMin = spreadsheet["F3"].value + 1
        rMax = spreadsheet["F4"].value + 1 
        loops = spreadsheet["F5"].value

        return rMin, rMax, loops

    def getTexts(self):
        spreadsheet = self.archive["Texts"]

        column = spreadsheet["D12"].value + 1
        row = spreadsheet["D13"].value + 1

        text = spreadsheet.cell(row=row, column=column).value

        return text

    def bingLoop(self):

        text = self.getTexts()
        rMin, rMax, loops = self.getAccounts()

        credentials = self.obtainEmails(rMin, rMax)

        for account, password in credentials.items():
            self.openChrome() 
            self.openBing(account, password)
            self.startSearch(text, loops)
            self.dayStreak()
            self.closeChrome()
            time.sleep(1)

bingPilot = BingPilot()